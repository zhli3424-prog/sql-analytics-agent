from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.orm import Session

from app.config import settings
from app.database import json_value
from app.models import Category, Customer, Order, OrderItem, Payment, Product, Refund


@dataclass(frozen=True)
class TableSpec:
    model: type
    columns: tuple[str, ...]
    converters: dict[str, Callable[[Any], Any]]
    optional: frozenset[str] = frozenset()


def integer(value: Any) -> int:
    if isinstance(value, bool):
        raise ValueError("必须是整数")
    number = int(str(value).strip())
    if number <= 0:
        raise ValueError("必须大于 0")
    return number


def nonnegative_decimal(value: Any) -> Decimal:
    try:
        number = Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError("必须是有效金额") from exc
    if not number.is_finite() or number < 0:
        raise ValueError("金额不能小于 0")
    return number


def text_value(value: Any) -> str:
    result = str(value).strip()
    if not result:
        raise ValueError("不能为空")
    return result


def date_value(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def datetime_value(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))


TABLE_SPECS = {
    "categories": TableSpec(Category, ("id", "name"), {"id": integer, "name": text_value}),
    "customers": TableSpec(
        Customer,
        ("id", "registered_at", "region", "channel"),
        {"id": integer, "registered_at": date_value, "region": text_value, "channel": text_value},
    ),
    "products": TableSpec(
        Product,
        ("id", "category_id", "name", "unit_cost", "list_price"),
        {
            "id": integer,
            "category_id": integer,
            "name": text_value,
            "unit_cost": nonnegative_decimal,
            "list_price": nonnegative_decimal,
        },
    ),
    "orders": TableSpec(
        Order,
        ("id", "customer_id", "ordered_at", "status", "region"),
        {
            "id": integer,
            "customer_id": integer,
            "ordered_at": datetime_value,
            "status": text_value,
            "region": text_value,
        },
    ),
    "order_items": TableSpec(
        OrderItem,
        ("id", "order_id", "product_id", "quantity", "unit_price", "discount_amount"),
        {
            "id": integer,
            "order_id": integer,
            "product_id": integer,
            "quantity": integer,
            "unit_price": nonnegative_decimal,
            "discount_amount": nonnegative_decimal,
        },
    ),
    "payments": TableSpec(
        Payment,
        ("id", "order_id", "method", "amount", "paid_at"),
        {
            "id": integer,
            "order_id": integer,
            "method": text_value,
            "amount": nonnegative_decimal,
            "paid_at": datetime_value,
        },
        frozenset({"paid_at"}),
    ),
    "refunds": TableSpec(
        Refund,
        ("id", "order_id", "amount", "reason", "refunded_at"),
        {
            "id": integer,
            "order_id": integer,
            "amount": nonnegative_decimal,
            "reason": text_value,
            "refunded_at": datetime_value,
        },
    ),
}


class ImportValidationError(ValueError):
    pass


def parse_import(filename: str, content: bytes) -> tuple[list[str], list[list[Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        headers, raw_rows = read_csv(content)
    elif suffix == ".xlsx":
        headers, raw_rows = read_xlsx(content)
    else:
        raise ImportValidationError("只支持 .csv 和 .xlsx 文件")
    return headers, raw_rows


def read_csv(content: bytes) -> tuple[list[str], list[list[Any]]]:
    decoded = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ImportValidationError("CSV 编码无法识别，请使用 UTF-8")
    reader = csv.reader(io.StringIO(decoded))
    rows = list(reader)
    if not rows:
        raise ImportValidationError("文件为空")
    return [str(value).strip() for value in rows[0]], rows[1:]


def read_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError("Excel 文件损坏或格式不受支持") from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        first = next(iterator, None)
        if first is None:
            raise ImportValidationError("文件为空")
        rows = []
        for index, row in enumerate(iterator, start=1):
            if index > settings.max_import_rows:
                raise ImportValidationError(f"单次最多导入 {settings.max_import_rows} 行")
            rows.append(list(row))
        return [str(value or "").strip() for value in first], rows
    finally:
        workbook.close()


def validate_rows(table_name: str, headers: list[str], raw_rows: list[list[Any]]) -> list[dict[str, Any]]:
    spec = TABLE_SPECS.get(table_name)
    if spec is None:
        raise ImportValidationError("目标表不允许导入")
    if len(headers) != len(set(headers)) or any(not header for header in headers):
        raise ImportValidationError("表头存在空值或重复字段")
    unknown = set(headers) - set(spec.columns)
    missing = set(spec.columns) - set(headers) - set(spec.optional)
    if unknown:
        raise ImportValidationError(f"存在未知字段：{', '.join(sorted(unknown))}")
    if missing:
        raise ImportValidationError(f"缺少必填字段：{', '.join(sorted(missing))}")

    rows: list[dict[str, Any]] = []
    errors: list[str] = []
    for row_number, values in enumerate(raw_rows, start=2):
        if not any(value not in (None, "") for value in values):
            continue
        if len(values) > len(headers):
            errors.append(f"第 {row_number} 行列数超过表头")
            continue
        raw = dict(zip(headers, values))
        converted: dict[str, Any] = {}
        for column in spec.columns:
            value = raw.get(column)
            if value in (None, "") and column in spec.optional:
                converted[column] = None
                continue
            try:
                converted[column] = spec.converters[column](value)
            except (TypeError, ValueError) as exc:
                errors.append(f"第 {row_number} 行 {column}：{exc}")
                break
        else:
            if table_name == "orders" and converted["status"] not in {"paid", "shipped", "completed", "cancelled"}:
                errors.append(f"第 {row_number} 行 status：不支持的订单状态")
                continue
            rows.append(converted)
        if len(errors) >= 20:
            break
    if errors:
        raise ImportValidationError("；".join(errors))
    if not rows:
        raise ImportValidationError("没有可导入的数据行")
    if len(rows) > settings.max_import_rows:
        raise ImportValidationError(f"单次最多导入 {settings.max_import_rows} 行")
    return rows


def preview_rows(rows: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    return [{key: json_value(value) for key, value in row.items()} for row in rows[:limit]]


def upsert_rows(session: Session, table_name: str, rows: list[dict[str, Any]]) -> None:
    spec = TABLE_SPECS[table_name]
    for offset in range(0, len(rows), 1000):
        statement = insert(spec.model).values(rows[offset : offset + 1000])
        updates = {
            column: getattr(statement.excluded, column)
            for column in spec.columns
            if column != "id"
        }
        session.execute(statement.on_conflict_do_update(index_elements=[spec.model.id], set_=updates))


def table_metadata() -> list[dict[str, Any]]:
    return [
        {"name": name, "columns": list(spec.columns), "optional": sorted(spec.optional)}
        for name, spec in TABLE_SPECS.items()
    ]
